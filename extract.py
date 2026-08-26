# -*- coding: utf-8 -*-
"""
ITS Master Excel -> JSON 추출 모듈
엑셀 구조가 바뀌어도 여기만 고치면 되도록 시트별로 분리
"""
import re
import openpyxl


def _s(v):
    """셀값 -> 정리된 문자열"""
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    if s in ("-", "–", "—", "#N/A", "N/A", "None", "nan"):
        return ""
    return s


def _ffill(rows, cols):
    """병합셀로 비어있는 컬럼을 위 행 값으로 채움"""
    last = {c: "" for c in cols}
    out = []
    for r in rows:
        r = list(r)
        for c in cols:
            if c < len(r) and _s(r[c]):
                last[c] = _s(r[c])
            elif c < len(r):
                r[c] = last[c]
        out.append(r)
    return out


# ─────────────────────────────────────────────
# 버전 비교 (권고역전 탐지용)
# ─────────────────────────────────────────────
VER_RE = re.compile(r"^\d+(?:\.\d+){1,5}$")

def clean_ver(v):
    """순수 버전 토큰만 통과. 제품명·설명 섞인 값은 None."""
    if not v:
        return None
    s = str(v).strip()
    if "," in s or "/" in s:        # 다중 후보 · 경로 -> 비교 불가
        return None
    s = re.sub(r"\s*\(.*?\)\s*", "", s).strip()   # 괄호 주석 제거
    if not VER_RE.match(s):
        return None
    if set(s.replace(".", "")) <= {"0"}:            # "0", "0.0" 등
        return None
    return s


def vparts(v):
    """'9.0.105' -> (9,0,105). 순수 버전이 아니면 None"""
    s = clean_ver(v)
    if not s:
        return None
    return tuple(int(x) for x in s.split(".")[:6])


def vcmp(a, b):
    """a>b:1, a==b:0, a<b:-1, 비교불가:None"""
    pa, pb = vparts(a), vparts(b)
    if not pa or not pb:
        return None
    # 자릿수 다르면 짧은 쪽 0 패딩
    n = max(len(pa), len(pb))
    pa = pa + (0,) * (n - len(pa))
    pb = pb + (0,) * (n - len(pb))
    return (pa > pb) - (pa < pb)


def same_line(a, b):
    """같은 메이저.마이너 라인인지 (라인 전환 vs 패치 상향 구분)"""
    pa, pb = vparts(a), vparts(b)
    if not pa or not pb:
        return False
    return pa[:2] == pb[:2]


# ─────────────────────────────────────────────
# 01_표준_권고버전마스터
# ─────────────────────────────────────────────
def sheet01(wb):
    ws = wb["01_표준_권고버전마스터"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    rows = _ffill(rows, [0, 1])  # OS기반, 계층 병합셀 채우기
    out = []
    for r in rows:
        base, layer, prod = _s(r[0]), _s(r[1]), _s(r[2])
        if not prod or prod.startswith("※"):
            continue
        out.append({
            "base":  base,          # AIX / Linux (RHEL·Rocky·OL) / Windows Server
            "layer": layer,         # DB / WAS / WEB / 미들웨어 / 필수SW
            "prod":  prod,
            "std":   _s(r[3]),      # 표준 지정버전
            "floor": _s(r[4]),      # 표준 하한(≥)
            "jdk":   _s(r[5]),
            "eos":   _s(r[6]),
            "note":  _s(r[7]),
        })
    return out


# ─────────────────────────────────────────────
# 02_자산RAWDATA_통합  (자산 마스터)
# ─────────────────────────────────────────────
ASSET_COLS = ["id", "layer", "env", "svrname", "sys", "os", "dual",
              "prod", "jdk", "cur", "rec", "eosdate", "eos", "upd",
              "reason", "memo"]

JUNK_ENV = {"DBMS아님", "운영구분"}
JUNK_LAYER = {"계층"}


def sheet02(wb):
    ws = wb["02_자산RAWDATA_통합"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    out = []
    dropped = []
    for r in rows:
        d = {ASSET_COLS[i]: _s(r[i]) if i < len(r) else "" for i in range(len(ASSET_COLS))}
        if not d["id"] or d["id"].startswith("※"):
            continue
        # 헤더 혼입 / 비대상 행 제거
        if d["layer"] in JUNK_LAYER or d["env"] in JUNK_ENV:
            dropped.append(d["id"])
            continue
        out.append(d)
    return out, dropped


def os_family(os_str):
    """OS 원문 -> AIX / Linux / Windows / 기타"""
    s = (os_str or "").lower()
    if "aix" in s:
        return "AIX"
    if any(k in s for k in ("windows", "win20", "win 20")):
        return "Windows"
    if any(k in s for k in ("redhat", "red hat", "rhel", "centos", "rocky",
                            "oracle linux", "ubuntu", "linux", "suse")):
        return "Linux"
    if "esxi" in s or "vmware" in s:
        return "가상화"
    if not s.strip():
        return "미기재"
    return "기타"


def os_short(os_str):
    """OS 원문 -> 짧은 표시명"""
    s = (os_str or "").strip()
    if not s:
        return ""
    fam = os_family(s)
    if fam == "AIX":
        m = re.search(r"(\d+\.\d+)", s)
        return f"AIX {m.group(1)}" if m else "AIX"
    if fam == "Windows":
        m = re.search(r"(2008|2012|2016|2019|2022|2025)", s)
        r2 = " R2" if "r2" in s.lower() else ""
        return f"Windows {m.group(1)}{r2}" if m else "Windows"
    if fam == "Linux":
        m = re.search(r"(\d+\.\d+)", s)
        v = m.group(1) if m else ""
        low = s.lower()
        if "centos" in low:
            return f"CentOS {v}".strip()
        if "rocky" in low:
            return f"Rocky {v}".strip()
        if "oracle linux" in low:
            return f"OL {v}".strip()
        if "ubuntu" in low:
            return f"Ubuntu {v}".strip()
        return f"RHEL {v}".strip()
    return s[:22]


# ─────────────────────────────────────────────
# 03_호환성_Matrix
# ─────────────────────────────────────────────
def sheet03(wb):
    ws = wb["03_호환성_Matrix"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    rows = _ffill(rows, [0, 1])
    out = []
    for r in rows:
        cat, prod, ver = _s(r[0]), _s(r[1]), _s(r[2])
        if not ver or ver.startswith("범례"):
            continue
        syms = [_s(r[i]) or "?" for i in range(3, 9)]
        out.append({
            "cat":  cat.replace("\n", " "),
            "prod": prod,
            "ver":  ver,
            "syms": syms,          # AIX7.1 AIX7.2 Lin8 Lin9 Win2019 Win2022
            "note": _s(r[9]),
        })
    return out


COMPAT_COLS = ["AIX 7.1", "AIX 7.2", "Linux 8.x", "Linux 9.x",
               "Win 2019", "Win 2022"]


# ─────────────────────────────────────────────
# 04_WAS_WEB_JDK요구
# ─────────────────────────────────────────────
def sheet04(wb):
    ws = wb["04_WAS_WEB_JDK요구"]
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        prod = _s(r[1])
        if not prod or prod.startswith("※"):
            continue
        out.append({
            "prod": prod,
            "min":  _s(r[2]),
            "rec":  _s(r[3]),
            "obs":  _s(r[4]),
            "note": _s(r[5]),
        })
    return out


# ─────────────────────────────────────────────
# 05_EOL_전환참고
# ─────────────────────────────────────────────
def sheet05(wb):
    ws = wb["05_EOL_전환참고"]
    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        prod = _s(r[1])
        if not prod or prod.startswith("※"):
            continue
        out.append({
            "prod":   prod,
            "ver":    _s(r[2]),
            "eos":    _s(r[3]),
            "state":  _s(r[4]),   # ✗ / △ / ○
            "action": _s(r[5]),
        })
    return out


# ─────────────────────────────────────────────
# 07_필수SW_OS별버전
# ─────────────────────────────────────────────
def sheet07(wb):
    ws = wb["07_필수SW_OS별버전"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    hdr = [_s(c) for c in rows[0]]
    sw_names = [h for h in hdr[7:] if h]
    out = []
    for r in rows[1:]:
        osn = _s(r[0])
        if not osn or osn.startswith("※"):
            continue
        out.append({
            "os":     osn,
            "ver":    _s(r[1]),
            "patch":  _s(r[2]),
            "state":  _s(r[3]),
            "eos":    _s(r[4]),
            "next":   _s(r[5]),
            "prio":   _s(r[6]),
            "sw":     [_s(r[7 + i]) for i in range(len(sw_names))],
        })
    return sw_names, out


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    std = sheet01(wb)
    assets, dropped = sheet02(wb)
    compat = sheet03(wb)
    jdk = sheet04(wb)
    eol = sheet05(wb)
    sw_names, sw = sheet07(wb)
    return dict(std=std, assets=assets, dropped=dropped, compat=compat,
                jdk=jdk, eol=eol, sw_names=sw_names, sw=sw)


if __name__ == "__main__":
    d = load("ITS_Master_v2_1.xlsx")
    for k in ("std", "assets", "compat", "jdk", "eol", "sw"):
        print(f"  {k}: {len(d[k])}")
    print(f"  dropped: {d['dropped']}")
    print(f"  sw_names: {d['sw_names']}")
    print("\n  std[0]:", d["std"][0])
    print("\n  assets[0]:", d["assets"][0])
    print("\n  compat[0]:", d["compat"][0])

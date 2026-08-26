# -*- coding: utf-8 -*-
"""
샘플 마스터파일 생성기

공개 저장소에서 실제 자산 데이터 없이 데모를 돌리기 위한 더미 데이터입니다.
업무명·자산ID는 모두 가공값이며 실제 시스템과 무관합니다.
제품명·버전·EOS 일자는 벤더 공개 정보를 사용합니다.

  python3 sample/make_sample.py
  → sample/ITS_Master_sample.xlsx
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(20260821)          # 매번 같은 결과가 나오도록 고정

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "ITS_Master_sample.xlsx")

G, GD, WH, GRAY = "FF00853F", "FF00612E", "FFFFFFFF", "FFF4F5F7"


def fill(c):
    return PatternFill("solid", fgColor=c)


def bd():
    s = Side(style="thin", color="FFC8CDD4")
    return Border(left=s, right=s, top=s, bottom=s)


def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color=G)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name="맑은 고딕", size=10, color="FF5C6660")


def header(ws, row, cols, widths=None):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(name="맑은 고딕", size=9, bold=True, color=WH)
        cell.fill = fill(G)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bd()
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def body(ws, start, rows):
    for ri, r in enumerate(rows, start=start):
        for ci, v in enumerate(r, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.fill = fill(GRAY if ri % 2 == 0 else "FFFFFFFF")
            cell.border = bd()
        ws.row_dimensions[ri].height = 17


# ══════════════════════════════════════════
# 참조 데이터 (벤더 공개 정보)
# ══════════════════════════════════════════
STD_ROWS = [
    # OS기반, 계층, 제품, 표준지정, 하한, JDK, EOS, 비고
    ("AIX", "DB",  "Oracle Database",       "19c (19.22)", "19.22", "",           "2027-04", "AIX는 7.2 TL1+ 필요. 12c 이하 전환대상"),
    ("",    "WAS", "IBM WebSphere",         "9.0.5.16",    "9.0.5.16", "Semeru 8/11", "8.5.5 2027", "AIX 최적. IBM SDK/Semeru 번들 사용"),
    ("",    "WAS", "TmaxSoft JEUS",         "8.5 Fix#10",  "8.5.0.10", "JDK 8/11", "8.5 2027", "국산 WAS. 라인 내 최신 Fix 유지"),
    ("",    "WEB", "IBM HTTP Server (IHS)", "9.0.5.16",    "9.0.5.16", "",         "8.5.5 2027", "WAS와 동일 라인 유지 권장"),
    ("Linux (RHEL·Rocky·OL)", "DB",  "Oracle Database", "19c (19.22)", "19.22", "",  "2027-04", "RHEL 8/9 인증. 전사 표준"),
    ("",    "DB",  "PostgreSQL",            "16.3",        "16.3",  "",           "2028-11", "AIX 미지원. RHEL 8/9 표준"),
    ("",    "DB",  "MariaDB",               "10.11 LTS",   "10.11", "",           "2028-02", "LTS 라인 유지"),
    ("",    "WAS", "Apache Tomcat",         "9.0.x (≥9.0.90)", "9.0.90", "JDK 8/11", "9.0 미정", "Java 이식성. 전사 표준"),
    ("",    "WAS", "Red Hat JBoss EAP",     "7.4.x",       "7.4.0", "JDK 11",     "2029-06", "EAP 8 전환 검토"),
    ("",    "WEB", "Apache HTTP Server",    "2.4.x (≥2.4.62)", "2.4.62", "",      "2.4 유지", "2.4 라인 내 최신 유지"),
    ("",    "WEB", "Nginx",                 "1.26.x (≥1.26.2)", "1.26.2", "",     "stable 유지", "mainline 라인 표준 제외"),
    ("Windows Server", "DB",  "MS SQL Server", "2019 (CU27)", "2019", "",         "2030-01", "2016 이하 전환대상"),
    ("",    "WAS", "Apache Tomcat",         "9.0.x (≥9.0.90)", "9.0.90", "JDK 8/11", "9.0 미정", "Windows Service 등록 확인"),
    ("",    "WEB", "Microsoft IIS",         "10.0",        "10.0",  "",           "OS 종속",  "OS 라이프사이클 따름"),
]

COMPAT_ROWS = [
    ("DB",  "Oracle Database",       "11g R2 (EOL)",   "△","△","✗","✗","✗","✗", "EOL(2020). AIX 레거시만 잔존 · 19c 전환대상"),
    ("",    "Oracle Database",       "19c (19.22) ★표준", "△","◎","◎","◎","◎","◎", "AIX는 7.2 TL1+ 필요. 전사 표준"),
    ("",    "MS SQL Server",         "2016 (EOL)",     "?","?","✗","✗","△","✗", "EOL 2026-07. 2019 전환"),
    ("",    "MS SQL Server",         "2019 (CU27) ★표준","?","?","○","△","◎","◎", "RHEL8 지원 · RHEL9 제한"),
    ("",    "PostgreSQL",            "16.x (16.3) ★표준","?","?","◎","◎","○","○", "AIX 미지원. RHEL8/9 표준"),
    ("WAS", "Apache Tomcat",         "8.5.x (EOL)",    "△","△","△","△","△","△", "라인 EOL 2024-03. 9.0/10.1 전환"),
    ("",    "Apache Tomcat",         "9.0.x ★표준",     "○","○","◎","◎","◎","◎", "Java 이식성. JDK8/11"),
    ("",    "Apache Tomcat",         "10.1.x ▲목표",    "○","○","◎","◎","◎","◎", "JDK11+ 필수. 차기 목표 라인"),
    ("",    "IBM WebSphere",         "8.5.5.x",        "◎","◎","○","△","○","△", "EOS 2027(권고). RHEL9 미인증"),
    ("",    "IBM WebSphere",         "9.0.5.x ▲목표",   "◎","◎","◎","○","◎","○", "전환 목표. Semeru JDK8/11"),
    ("",    "TmaxSoft JEUS",         "8.5 ★표준",       "◎","◎","◎","○","○","○", "국산 WAS. 라인 내 Fix 유지"),
    ("WEB", "Apache HTTP Server",    "2.4.x ★표준",     "○","○","◎","◎","○","○", "2.4 라인 내 최신 유지"),
    ("",    "Nginx",                 "1.26.x ★표준",    "△","△","◎","◎","○","○", "stable 라인만 표준"),
    ("",    "IBM HTTP Server (IHS)", "9.0.5.x ▲목표",   "◎","◎","○","△","○","△", "WAS와 라인 일치 권장"),
]

JDK_ROWS = [
    ("Apache Tomcat 9.0",   "JDK 8",      "JDK 11",     "1.8.0_382", "8 최소 · 11 권고"),
    ("Apache Tomcat 10.1",  "JDK 11",     "JDK 17",     "11.0.22",   "JDK 11 미만 기동 불가"),
    ("IBM WebSphere 8.5.5", "IBM SDK 8",  "Semeru 8",   "1.8.0_351", "번들 SDK 사용 강제"),
    ("IBM WebSphere 9.0.5", "Semeru 8",   "Semeru 11",  "11.0.20",   "IBM Semeru 권장"),
    ("TmaxSoft JEUS 8.5",   "JDK 8",      "JDK 11",     "1.8.0_362", "Tmax 인증 JDK 확인"),
    ("Red Hat JBoss EAP 7.4", "JDK 8",    "JDK 11",     "11.0.21",   "EAP 8은 JDK 17"),
]

EOL_ROWS = [
    ("Oracle Database",  "11g R2",   "2020-12-31", "✗", "19c 전환 (Premier 2027-04)"),
    ("Oracle Database",  "19c",      "2027-04-30", "○", "RU 최신 유지"),
    ("MS SQL Server",    "2016",     "2026-07-14", "△", "2019 이상 전환"),
    ("MS SQL Server",    "2019",     "2030-01-08", "○", "CU 최신 유지"),
    ("Apache Tomcat",    "8.5.x",    "2024-03-31", "✗", "9.0 / 10.1 전환"),
    ("Apache Tomcat",    "9.0.x",    "미정",        "○", "라인 내 최신 유지"),
    ("IBM WebSphere",    "8.5.5",    "2027-04-30", "△", "9.0.5 / Liberty 전환"),
    ("PostgreSQL",       "12.x",     "2024-11-14", "✗", "16.x 전환"),
    ("Nginx",            "1.24.x",   "2025-05-31", "△", "1.26 stable 전환"),
    ("AIX",              "7.1",      "2023-04-30", "✗", "7.2 TL5 / 7.3 전환"),
    ("AIX",              "7.2",      "2026-04-30", "△", "7.3 전환 검토"),
    ("Red Hat Enterprise Linux", "7.9", "2024-06-30", "✗", "RHEL 8/9 전환"),
    ("Windows Server",   "2012 R2",  "2023-10-10", "✗", "2019 / 2022 전환"),
]

SW_NAMES = ["백신", "백업", "모니터링", "접근제어", "형상관리", "로그수집", "APM"]
SW_ROWS = [
    ("AIX", "7.1", "TL05", "단종(EOS)", "2023-04-30", "AIX 7.3 TL1", "1순위 (즉시)",
     "8.3", "7.65", "6.1.12", "-", "3.0.5074", "2.0.31", "-"),
    ("AIX", "7.2", "TL05", "지원중", "2026-04-30", "AIX 7.3 TL1", "2순위 (2주)",
     "8.5", "8.2.4", "7.3.4", "-", "3.0.5074", "4.0.5", "-"),
    ("Red Hat Enterprise Linux", "7.9", "-", "단종(EOS)", "2024-06-30", "RHEL 8.10", "1순위 (즉시)",
     "8.3", "8.1.1.6", "7.3.4", "2.1.0", "3.0.5074", "2.0.31", "5.5.4"),
    ("Red Hat Enterprise Linux", "8.10", "-", "지원중", "2029-05-31", "-", "4순위 (반기)",
     "8.5", "8.2.4", "7.3.4", "2.4.0", "3.0.5074", "4.0.5", "6.1.0"),
    ("Red Hat Enterprise Linux", "9.4", "-", "지원중", "2032-05-31", "-", "4순위 (반기)",
     "8.5", "8.2.4", "7.3.4", "2.4.0", "3.0.5074", "4.0.5", "6.1.0"),
    ("Windows Server", "2012 R2", "-", "단종(EOS)", "2023-10-10", "Windows Server 2022", "1순위 (즉시)",
     "8.3", "7.65", "6.1.12", "2.1.0", "3.0.5074", "2.0.31", "-"),
    ("Windows Server", "2019", "-", "지원중", "2029-01-09", "-", "3순위 (분기)",
     "8.5", "8.2.4", "7.3.4", "2.4.0", "3.0.5074", "4.0.5", "6.1.0"),
    ("Windows Server", "2022", "-", "지원중", "2031-10-14", "-", "4순위 (반기)",
     "8.5", "8.2.4", "7.3.4", "2.4.0", "3.0.5074", "4.0.5", "6.1.0"),
]

# ── 자산 생성용 (모두 가공값) ──
SYSTEMS = [
    "대외연계", "통합포털", "고객관리", "청구지급", "계약관리", "채널통합",
    "상품관리", "정산배치", "통합인증", "모바일서비스", "리포팅", "데이터허브",
    "알림발송", "문서관리", "이력추적", "코드관리", "권한관리", "배치스케줄러",
]
ENVS = ["운영", "개발", "테스트", "재해복구"]
DUAL = ["이중화", "단독", ""]

OS_POOL = [
    ("AIX 7.1 TL05 SP03", "AIX"),
    ("AIX 7.2 TL05 SP07", "AIX"),
    ("Red Hat Enterprise Linux 7.9", "Linux"),
    ("Red Hat Enterprise Linux 8.10", "Linux"),
    ("Red Hat Enterprise Linux 9.4", "Linux"),
    ("Windows Server 2012 R2", "Windows"),
    ("Windows Server 2019", "Windows"),
    ("Windows Server 2022", "Windows"),
]

# (계층, 제품, [(현재, 권고, EOS여부, 업데이트, 사유)])
PRODUCTS = {
    "WAS": [
        ("Tomcat", [
            ("8.5.100", "9.0.90, 10.1.24", "O", "불가", "8.5.x 라인 EOL(2024-03). 상위 라인 전환 필요"),
            ("9.0.57",  "9.0.90", "", "가능", "라인 내 최신 패치 상향 가능"),
            ("9.0.90",  "9.0.90", "", "불필요", ""),
            ("10.1.16", "10.1.24", "", "가능", "라인 내 최신 패치 상향 가능"),
        ]),
        ("WebSphere", [
            ("8.5.5.20", "8.5.5.26, 9.0.5.16", "", "가능", "Fix Pack 상향 또는 9.0 라인 전환"),
            ("9.0.5.16", "9.0.5.16", "", "불필요", ""),
            ("7.0.0.33", "9.0.5.16", "O", "불가", "7.0 라인 EOL. 9.0 전환 필요"),
        ]),
        ("JEUS", [
            ("8.5.0.0",  "8.5.0.10", "", "가능", "Fix 상향 가능"),
            ("8.5.0.10", "8.5.0.10", "", "불필요", ""),
        ]),
        ("JBoss EAP", [
            ("7.4.0", "7.4.14", "", "가능", "라인 내 최신 패치 상향 가능"),
        ]),
    ],
    "WEB": [
        ("Apache", [
            ("2.4.37", "2.4.62", "", "가능", "라인 내 최신 패치 상향 가능"),
            ("2.4.62", "2.4.62", "", "불필요", ""),
        ]),
        ("Nginx", [
            ("1.20.1", "1.26.2", "O", "불가", "1.20 라인 지원 종료. stable 라인 전환"),
            ("1.26.2", "1.26.2", "", "불필요", ""),
        ]),
        ("IHS", [
            ("8.5.5.20", "8.5.5.26, 9.0.5.16", "", "가능", "Fix Pack 상향 또는 9.0 라인 전환"),
            ("7.0.0.33", "9.0.5.16", "O", "불가", "7.0 라인 EOL. 9.0 전환 필요"),
        ]),
        ("IIS", [
            ("10.0", "10.0", "", "불필요", "OS 라이프사이클 따름"),
        ]),
    ],
    "DB": [
        ("Oracle (Enterprise)", [
            ("11.2.0.4", "19.22", "O", "불가", "11g EOL(2020-12). 19c 전환 필요"),
            ("19.15",    "19.22", "", "가능", "RU 상향 가능"),
            ("19.22",    "19.22", "", "불필요", ""),
        ]),
        ("MS-SQL", [
            ("2016 SP3", "2019 CU27", "O", "불가", "2016 EOL 2026-07. 2019 이상 전환"),
            ("2019 CU20", "2019 CU27", "", "가능", "CU 상향 가능"),
        ]),
        ("PostgreSQL", [
            ("12.14", "16.3", "O", "불가", "12.x EOL(2024-11). 16.x 전환"),
            ("16.3",  "16.3", "", "불필요", ""),
        ]),
        ("MariaDB", [
            ("10.6.8",  "10.11", "", "가능", "LTS 라인 상향 가능"),
            ("10.11.6", "10.11", "", "불필요", ""),
        ]),
    ],
}

JDKS = ["1.8.0_202", "1.8.0_382", "11.0.22", "17.0.10", ""]


def make_assets(n_per_layer=45):
    rows = []
    for layer in ("WAS", "WEB", "DB"):
        seq = 1
        for _ in range(n_per_layer):
            prod, variants = random.choice(PRODUCTS[layer])
            cur, rec, eos, upd, reason = random.choice(variants)
            os_full, fam = random.choice(OS_POOL)
            # AIX 에 IIS/MSSQL 같은 조합은 피함
            if fam == "AIX" and prod in ("IIS", "MS-SQL", "PostgreSQL", "MariaDB", "Nginx"):
                os_full, fam = random.choice([o for o in OS_POOL if o[1] == "Linux"])
            if fam == "Windows" and prod in ("JEUS", "IHS", "WebSphere"):
                os_full, fam = random.choice([o for o in OS_POOL if o[1] == "Linux"])
            jdk = random.choice(JDKS) if layer == "WAS" else ""
            rows.append([
                f"{layer}-{seq:03d}",
                layer,
                random.choice(ENVS),
                "",                                   # 서버명 — 비식별
                random.choice(SYSTEMS),
                os_full,
                random.choice(DUAL),
                prod,
                jdk,
                cur, rec,
                "", eos, upd, reason, "",
            ])
            seq += 1
    return rows


def main():
    wb = Workbook()

    # 00_안내
    ws = wb.active
    ws.title = "00_안내"
    title(ws, "인프라 표준화 마스터파일 — 샘플",
          "공개 저장소 데모용 더미 데이터입니다. 실제 자산과 무관합니다.")
    notes = [
        ["구분", "내용"],
        ["용도", "포털 빌드·기능 확인용 샘플"],
        ["자산", "가공된 업무명·자산ID (실제 시스템 아님)"],
        ["서버명", "비식별 원칙에 따라 공란"],
        ["제품·버전", "벤더 공개 정보 기반"],
        ["실제 운영", "사내 마스터파일로 교체해 사용하십시오"],
    ]
    header(ws, 4, notes[0], [16, 60])
    body(ws, 5, notes[1:])

    # 01_표준
    ws = wb.create_sheet("01_표준_권고버전마스터")
    title(ws, "OS 기반별 표준 권고버전 마스터", "AIX · Linux · Windows")
    header(ws, 4,
           ["OS 기반", "계층", "제품", "표준 지정버전", "표준 하한(≥)",
            "권고 JDK", "EOS/EOL(참고)", "OS별 제약 · 비고"],
           [22, 8, 24, 18, 14, 14, 14, 44])
    body(ws, 5, STD_ROWS)

    # 02_자산
    ws = wb.create_sheet("02_자산RAWDATA_통합")
    title(ws, "자산 원장 (샘플)", "서버명은 비식별 처리되어 공란입니다.")
    header(ws, 3,
           ["자산ID", "계층", "운영구분", "서버명", "업무/시스템", "OS", "이중화",
            "제품·종류", "JDK버전", "현재버전", "권고버전", "EOS일시", "EOS여부",
            "업데이트가능", "사유", "비고"],
           [10, 7, 9, 10, 16, 30, 8, 16, 13, 14, 20, 12, 8, 10, 40, 10])
    assets = make_assets()
    body(ws, 4, assets)

    # 03_호환성
    ws = wb.create_sheet("03_호환성_Matrix")
    title(ws, "제품 × OS 호환성 매트릭스", "◎ 인증 · ○ 지원 · △ 조건부 · ✗ 미지원 · ? 확인필요")
    header(ws, 4,
           ["분류", "제품", "버전", "AIX 7.1", "AIX 7.2", "Linux 8.x",
            "Linux 9.x", "Win 2019", "Win 2022", "제약 · 비고"],
           [8, 24, 20, 9, 9, 10, 10, 10, 10, 42])
    body(ws, 5, COMPAT_ROWS)

    # 04_JDK
    ws = wb.create_sheet("04_WAS_WEB_JDK요구")
    title(ws, "WAS · WEB 요구 JDK", "JDK 불일치는 미들웨어 장애의 주요 원인입니다.")
    header(ws, 3, ["", "제품 · 버전", "최소 JDK", "권고 JDK", "운영 JDK(관측)", "비고"],
           [4, 26, 14, 14, 16, 30])
    body(ws, 4, [("",) + r for r in JDK_ROWS])

    # 05_EOL
    ws = wb.create_sheet("05_EOL_전환참고")
    title(ws, "EOL · EOS 전환 참고", "✗ 종료 · △ 임박 · ○ 지원중")
    header(ws, 3, ["", "제품", "버전", "EOS", "상태", "전환 권고"],
           [4, 28, 14, 14, 8, 34])
    body(ws, 4, [("",) + r for r in EOL_ROWS])

    # 06_판정규칙
    ws = wb.create_sheet("06_범례_판정규칙")
    title(ws, "판정 규칙 · 범례")
    header(ws, 3, ["규칙", "대상", "내용", "효과", "주기"], [10, 10, 52, 16, 10])
    body(ws, 4, [
        ("R-C01", "공통", "해당 라인의 최신 안정화 패치 레벨을 권고 하한으로 지정", "표준 후보", "분기"),
        ("R-C02", "공통", "EOS·연장지원 종료 라인은 표준 제외, 전환 대상 분류", "미충족 시 금지", "상시"),
        ("R-C03", "공통", "미조치 CVSS 9.0 이상 취약점 없음", "미충족 시 금지", "상시"),
        ("R-W02", "WAS", "JDK 최소·권고 버전 병기", "결과 표기 형식", "상시"),
        ("R-R01", "릴리스", "롤링/mainline 라인은 표준 대상 제외", "대상 제외", "연1회"),
        ("R-V01", "공통", "출처 URL·확인일 없는 항목은 보류로 두고 표준 미포함", "보류 처리", "상시"),
        ("R-V02", "공통", "권고 하한이 관측 현재버전보다 낮으면 역전으로 표시하고 상향", "역전 경고", "분기"),
    ])

    # 07_필수SW
    ws = wb.create_sheet("07_필수SW_OS별버전")
    title(ws, "OS별 필수 소프트웨어 버전", "AI 검색 불가 항목 — 담당자 수동 확인 대상")
    header(ws, 4,
           ["OS명", "버전", "패치", "상태", "EOS날짜", "후속 권고", "우선순위"] + SW_NAMES,
           [24, 10, 8, 12, 12, 20, 14] + [11] * len(SW_NAMES))
    body(ws, 5, SW_ROWS)

    wb.save(OUT)
    print(f"생성 완료 · {OUT}")
    print(f"  자산 {len(assets)}건 · 표준 {len(STD_ROWS)}행 · 호환성 {len(COMPAT_ROWS)}행")
    print(f"  크기 {os.path.getsize(OUT):,} bytes")


if __name__ == "__main__":
    main()
